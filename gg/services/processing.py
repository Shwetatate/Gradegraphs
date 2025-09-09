import io
import re
import pandas as pd
from openpyxl import load_workbook

from .utils import extract_max_marks_from_header, is_practical_column


def find_data_start_row(file_obj, sheet_name=0):
    try:
        wb = load_workbook(file_obj)
        ws = wb.active if sheet_name == 0 else wb[sheet_name]

        for row in range(1, 31):
            row_values = []
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    row_values.append(str(cell_value).lower())

            row_str = ' '.join(row_values)

            if any(keyword in row_str for keyword in ['sr.no', 'sr no', 'serial', 's.no', 'roll', 'student', 'ise', 'mse', 'ese']):
                print(f"Found data start indicators at row {row}: {row_str}")
                return row - 1

        print("Data start not found, assuming row 7")
        return 7
    except Exception as e:
        print(f"Error finding data start row: {e}")
        return 7


def extract_subject_info(file_obj, sheet_name=0):
    try:
        wb = load_workbook(file_obj)
        ws = wb.active if sheet_name == 0 else wb[sheet_name]

        subjects = {}

        subject_patterns = [
            r'([A-Z]{2,3}\d{3,4})\s*:\s*(.+)',
            r'([A-Z]{2,3}\d{3,4})\s+(.+)',
            r'([A-Z]{2,3}\d{3,4})(.+)',
        ]

        for row in range(1, 16):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 5:
                    for pattern in subject_patterns:
                        match = re.search(pattern, cell_value)
                        if match:
                            subject_code = match.group(1)
                            subject_name = match.group(2).strip()
                            subjects[subject_code] = subject_name
                            print(f"Found subject: {subject_code} -> {subject_name}")

        print(f"Total subjects found: {len(subjects)}")
        return subjects
    except Exception as e:
        print(f"Error extracting subject info: {e}")
        return {}


def safe_get_value(df, row_idx, col_name):
    try:
        if row_idx >= len(df):
            return None
        if col_name not in df.columns:
            return None
        col_idx = df.columns.get_loc(col_name)
        if isinstance(col_idx, (list,)):
            col_idx = col_idx[0]
        elif isinstance(col_idx, slice):
            col_idx = col_idx.start
        value = df.iloc[row_idx, col_idx]
        if hasattr(value, 'item'):
            return value.item()
        return value
    except Exception as e:
        print(f"Error getting value at row {row_idx}, col {col_name}: {e}")
        return None


def is_valid_number(value):
    if value is None:
        return False
    try:
        float_val = float(value)
        import numpy as np
        return not (np.isnan(float_val) or np.isinf(float_val))
    except (ValueError, TypeError):
        return False


def calculate_performance_metrics(df):
    print("\n" + "="*60)
    print("CALCULATING PERFORMANCE METRICS (GRADE GRAPH)")
    print("="*60)

    student_info_cols = ['Sr_No', 'Roll_No', 'Student_Name']
    all_subject_cols = [col for col in df.columns if col not in student_info_cols]

    print("\n🔍 ANALYZING COLUMNS FOR PRACTICAL DETECTION:")
    practical_cols = []
    for col in all_subject_cols:
        if is_practical_column(col):
            practical_cols.append(col)

    theory_cols = [col for col in all_subject_cols if col not in practical_cols and any(keyword in str(col).upper() for keyword in ['ISE', 'MSE', 'ESE'])]

    print(f"\n📊 COLUMN CLASSIFICATION RESULTS:")
    print(f"Found {len(theory_cols)} theory columns and {len(practical_cols)} practical columns")
    print(f"Theory columns: {theory_cols[:5]}{'...' if len(theory_cols) > 5 else ''}")
    print(f"Practical columns: {practical_cols}")

    def extract_max_marks(col_name):
        marks = extract_max_marks_from_header(col_name)
        print(f"    Using detected max marks: {col_name} -> {marks}")
        return marks

    print(f"\n1️⃣ Calculating Academic Performance %...")
    if all_subject_cols:
        academic_performance_list = []
        for row_idx in range(len(df)):
            total_obtained = 0
            total_maximum = 0
            for col in all_subject_cols:
                obtained_marks = safe_get_value(df, row_idx, col)
                max_marks = extract_max_marks(col)
                if is_valid_number(obtained_marks):
                    total_obtained += float(obtained_marks)
                    total_maximum += max_marks
            if total_maximum > 0:
                percentage = (total_obtained / total_maximum) * 100
                academic_performance_list.append(round(percentage, 2))
            else:
                academic_performance_list.append(0)
        df['Academic_Performance_%'] = academic_performance_list
        print(f"✅ Academic Performance calculated using sum method for {len(df)} students")
        if len(df) > 0:
            first_student_obtained = 0
            first_student_max = 0
            for col in all_subject_cols:
                val = safe_get_value(df, 0, col)
                if is_valid_number(val):
                    first_student_obtained += float(val)
                    first_student_max += extract_max_marks(col)
            if first_student_max > 0:
                print(f"📊 Sample calculation for first student:")
                print(f"    Total obtained: {first_student_obtained}")
                print(f"    Total maximum: {first_student_max}")
                print(f"    Percentage: {(first_student_obtained/first_student_max)*100:.2f}%")
    else:
        print("⚠️ Warning: No subject columns found for Academic Performance calculation")
        df['Academic_Performance_%'] = 0

    print(f"\n2️⃣ Calculating Previous Performance Analysis...")
    def get_previous_performance_analysis(academic_perf):
        if pd.isna(academic_perf) or academic_perf is None:
            return "weak"
        try:
            perf_val = float(academic_perf)
            if perf_val < 70:
                return "weak"
            elif perf_val < 90:
                return "ok"
            else:
                return "bright"
        except (ValueError, TypeError):
            return "weak"

    df['Previous_Performance_Analysis'] = df['Academic_Performance_%'].apply(get_previous_performance_analysis)
    perf_dist = df['Previous_Performance_Analysis'].value_counts()
    print(f"📊 Previous Performance Analysis distribution: {dict(perf_dist)}")

    print(f"\n3️⃣ Calculating Practical % (Only identified PRACTICAL/PR columns)...")
    if practical_cols:
        practical_performance_list = []
        print(f"🔍 Practical calculation details:")
        total_practical_max_possible = 0
        for col in practical_cols:
            max_marks = extract_max_marks(col)
            total_practical_max_possible += max_marks
            print(f"  ✅ {col}: Max marks = {max_marks}")
        print(f"📊 Total maximum practical marks possible: {total_practical_max_possible}")
        for row_idx in range(len(df)):
            practical_obtained = 0
            practical_maximum = 0
            for col in practical_cols:
                obtained_marks = safe_get_value(df, row_idx, col)
                max_marks = extract_max_marks(col)
                if is_valid_number(obtained_marks):
                    practical_obtained += float(obtained_marks)
                    practical_maximum += max_marks
            if practical_maximum > 0:
                percentage = (practical_obtained / practical_maximum) * 100
                practical_performance_list.append(round(percentage, 2))
            else:
                practical_performance_list.append(0)
        df['Practical_%'] = practical_performance_list
        print(f"✅ Practical Performance calculated for {len(df)} students using identified PRACTICAL/PR columns only")
        if len(df) > 0:
            first_student_practical = 0
            first_student_practical_max = 0
            practical_details = []
            for col in practical_cols:
                val = safe_get_value(df, 0, col)
                if is_valid_number(val):
                    obtained = float(val)
                    max_marks = extract_max_marks(col)
                    first_student_practical += obtained
                    first_student_practical_max += max_marks
                    practical_details.append(f"{col}: {obtained}/{max_marks}")
            if first_student_practical_max > 0:
                print(f"📊 Sample practical calculation for first student:")
                print(f"    Practical breakdown: {'; '.join(practical_details)}")
                print(f"    Total practical obtained: {first_student_practical}")
                print(f"    Total practical maximum: {first_student_practical_max}")
                print(f"    Practical percentage: {(first_student_practical/first_student_practical_max)*100:.2f}%")
        excluded_cols = [col for col in all_subject_cols if not is_practical_column(col) and any(keyword in str(col).upper() for keyword in ['TW', 'ENVIRONMENTAL', 'ENGINEERING']) and 'PR' not in str(col).upper()]
        if excluded_cols:
            print(f"❌ Columns EXCLUDED from Practical % calculation:")
            for col in excluded_cols:
                print(f"    • {col}")
    else:
        print("⚠️ Warning: No PRACTICAL/PR columns found. Setting Practical % to 0")
        df['Practical_%'] = 0

    print(f"\n4️⃣ Calculating Coding Expertise...")
    def get_coding_expertise(practical_perf):
        if pd.isna(practical_perf) or practical_perf is None:
            return "B"
        try:
            perf_val = float(practical_perf)
            if perf_val < 65:
                return "B"
            elif perf_val <= 90:
                return "I"
            else:
                return "A"
        except (ValueError, TypeError):
            return "B"

    df['Coding_Expertise'] = df['Practical_%'].apply(get_coding_expertise)
    coding_dist = df['Coding_Expertise'].value_counts()
    print(f"📊 Coding Expertise distribution: {dict(coding_dist)}")

    print(f"\n5️⃣ Calculating Performance Analysis using Grade Graph Rules...")
    def get_performance_analysis_grade_graph(prev_perf, coding_exp):
        if pd.isna(prev_perf) or pd.isna(coding_exp) or prev_perf is None or coding_exp is None:
            return "weak"
        prev_perf = str(prev_perf).strip().lower()
        coding_exp = str(coding_exp).strip().upper()
        if prev_perf == "weak":
            if coding_exp == "I":
                return "weak"
            elif coding_exp == "A":
                return "ok"
            elif coding_exp == "B":
                return "weak"
            else:
                return "weak"
        elif prev_perf == "ok":
            if coding_exp == "I":
                return "ok"
            elif coding_exp == "A":
                return "ok"
            elif coding_exp == "B":
                return "ok"
            else:
                return "ok"
        elif prev_perf == "bright":
            if coding_exp == "I":
                return "bright"
            elif coding_exp == "A":
                return "bright"
            elif coding_exp == "B":
                return "ok"
            else:
                return "bright"
        else:
            return "weak"

    df['Performance_Analysis'] = df.apply(lambda row: get_performance_analysis_grade_graph(
        row['Previous_Performance_Analysis'],
        row['Coding_Expertise']
    ), axis=1)

    final_dist = df['Performance_Analysis'].value_counts()
    print(f"📊 Final Performance Analysis distribution (Grade Graph): {dict(final_dist)}")

    print("\n" + "="*60)
    print("✅ PERFORMANCE METRICS CALCULATION COMPLETE (GRADE GRAPH)")
    print("="*60)

    return df


def process_excel_file(uploaded_file):
    try:
        file_obj = io.BytesIO(uploaded_file.getvalue())

        data_start_row = find_data_start_row(file_obj, sheet_name=0)
        print(f"Data starts at row: {data_start_row}")

        file_obj.seek(0)
        wb = load_workbook(file_obj)
        ws = wb.active

        file_obj.seek(0)
        subjects_info = extract_subject_info(file_obj, sheet_name=0)
        print(f"Extracted subjects: {subjects_info}")

        header_rows = []
        for row in range(data_start_row + 1, data_start_row + 4):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value if cell_value is not None else "")
            header_rows.append(row_data)

        print(f"Header rows: {header_rows}")

        file_obj.seek(0)
        df = pd.read_excel(file_obj, sheet_name=0, skiprows=data_start_row + 3, header=None)
        df = df.dropna(axis=1, how='all')
        df = df.dropna(axis=0, how='all')
        print(f"DataFrame shape after reading: {df.shape}")
        print(f"First few rows of raw data:")
        print(df.head(2))

        new_columns = []
        subject_row = header_rows[0] if len(header_rows) > 0 else []
        marks_row = header_rows[1] if len(header_rows) > 1 else []
        assessment_row = header_rows[2] if len(header_rows) > 2 else []

        print(f"Processing {len(df.columns)} columns...")

        current_subject = ""
        for i in range(len(df.columns)):
            subject_info = subject_row[i] if i < len(subject_row) else ""
            assessment_info = assessment_row[i] if i < len(assessment_row) else ""
            marks_info = marks_row[i] if i < len(marks_row) else ""

            subject_text = str(subject_info).strip() if subject_info is not None else ""
            assessment_text = str(assessment_info).strip() if assessment_info is not None else ""
            marks_text = str(marks_info).strip() if marks_info is not None else ""

            print(f"Column {i}: Subject='{subject_text}' | Assessment='{assessment_text}' | Marks='{marks_text}'")

            if i == 0:
                col_name = "Sr_No"
            elif i == 1:
                col_name = "Roll_No"
            elif i == 2:
                col_name = "Student_Name"
            else:
                subject_text_lower = subject_text.lower()
                assessment_text_lower = assessment_text.lower()
                if subject_text and subject_text not in ["", "nan", "None"]:
                    subject_code_pattern = r'([A-Z]{2,3}\d{3,4})'
                    has_subject_code = re.search(subject_code_pattern, subject_text)
                    is_meaningful_text = len(subject_text.strip()) > 3 and not subject_text.strip().isdigit()
                    if has_subject_code or is_meaningful_text:
                        if ":" in subject_text:
                            current_subject = subject_text.split(":", 1)[1].strip().title()
                        else:
                            current_subject = subject_text.strip().title()
                        current_subject = current_subject.replace("&", "and")
                        current_subject = re.sub(r'\s+', ' ', current_subject).strip()
                        print(f"  -> Detected new subject: '{current_subject}'")
                assessment_type = assessment_text.upper() if assessment_text else ""
                marks_info = marks_text if marks_text else ""
                if current_subject and assessment_type and assessment_type in ['ISE', 'MSE', 'ESE', 'PRACTICAL', 'TW', 'PR']:
                    if marks_info and marks_info.isdigit():
                        col_name = f"{current_subject} {assessment_type} ({marks_info})"
                    else:
                        col_name = f"{current_subject} {assessment_type}"
                    print(f"  -> Created: {col_name}")
                elif current_subject and assessment_type:
                    if marks_info and marks_info.isdigit():
                        col_name = f"{current_subject} {assessment_type} ({marks_info})"
                    else:
                        col_name = f"{current_subject} {assessment_type}"
                    print(f"  -> Created: {col_name}")
                else:
                    col_name = f"Column_{i}"
                    print(f"  -> Default: {col_name}")

            new_columns.append(col_name)

        df.columns = new_columns[:len(df.columns)]

        print(f"Column mapping complete. New columns:")
        for i, col in enumerate(df.columns):
            print(f"  {i}: {col}")

        if 'Sr_No' in df.columns:
            initial_rows = len(df)
            df = df[df['Sr_No'].notna()]
            df = df[df['Sr_No'] != '']
            df = df[df['Sr_No'] != 0]
            print(f"Removed {initial_rows - len(df)} rows with empty Sr_No")

        df = df.reset_index(drop=True)
        print(f"Data after cleanup: {df.shape}")

        non_numeric_cols = ['Sr_No', 'Roll_No', 'Student_Name']
        for col in df.columns:
            if col in non_numeric_cols:
                continue
            try:
                if col in df.columns and len(df[col]) > 0:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                    print(f"Converted column '{col}' to numeric")
            except Exception as e:
                print(f"Warning: Could not convert column '{col}' to numeric: {e}")

        df = calculate_performance_metrics(df)

        rename_map = {}
        for col in df.columns:
            if col == 'Sr_No':
                rename_map[col] = 'SR.No'
            elif col == 'Roll_No':
                rename_map[col] = 'Roll No'
            elif col == 'Student_Name':
                rename_map[col] = 'Name'

        df.rename(columns=rename_map, inplace=True)

        print("📊 Overall Average calculation REMOVED as per Grade Graph requirements")

        def classify_student_grade_graph(row):
            performance_analysis = row.get('Performance_Analysis', 'weak')
            if pd.notna(performance_analysis):
                perf_str = str(performance_analysis).strip().lower()
                if perf_str in ['weak', 'w']:
                    return "Weak"
                elif perf_str in ['ok', 'average']:
                    return "Average"
                elif perf_str in ['bright', 'b']:
                    return "Bright"
                else:
                    return "Unknown"
            else:
                return "Unknown"

        df['Category'] = df.apply(classify_student_grade_graph, axis=1)
        category_dist = df['Category'].value_counts()
        print(f"📊 Student Classification Results (Grade Graph): {dict(category_dist)}")

        suggestion_columns = ['SR.No', 'Roll No', 'Name', 'Category']
        performance_cols = ['Academic_Performance_%', 'Previous_Performance_Analysis', 'Practical_%', 'Coding_Expertise', 'Performance_Analysis']
        for col in performance_cols:
            if col in df.columns:
                suggestion_columns.insert(-1, col)
        available_cols = [col for col in suggestion_columns if col in df.columns]
        df_suggestions = df[available_cols].copy()

        print(f"Processing complete. Final shape: {df.shape}")
        print(f"✅ Enhanced analysis columns added (Grade Graph):")
        analysis_cols = ['Academic_Performance_%', 'Previous_Performance_Analysis', 'Practical_%', 'Coding_Expertise', 'Performance_Analysis']
        for col in analysis_cols:
            if col in df.columns:
                print(f"  ✓ {col}")

        return df, df_suggestions
    except Exception as e:
        print(f"Error processing file: {e}")
        import traceback
        traceback.print_exc()
        raise e


